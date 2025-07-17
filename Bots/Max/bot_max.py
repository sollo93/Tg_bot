import asyncio
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from os.path import exists
from datetime import datetime, timedelta, time as dt_time

from maxgram import MaxBot, MaxDispatcher, types

API_TOKEN = "ВАШ_MAX_ТОКЕН"
EXCEL_FILE = "appointments.xlsx"

bot = MaxBot(token=API_TOKEN)
dp = MaxDispatcher()

# --- Данные
specialists = ["Соловьева", "Соловьев"]
massage_types = ["Классический", "Точечный", "Спортивный", "Расслабляющий", "Антицеллюлитный"]
work_start = dt_time(8, 0); work_end = dt_time(23, 0)
lunch_start = dt_time(13, 0); lunch_end = dt_time(14, 0)

specialist_colors = {"Соловьева": "6272A4", "Соловьев": "50FA7B"}
specialist_user_ids = {"Соловьева": 1152741301, "Соловьев": 812829316}

def load_appointments():
    if exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
        return df.to_dict(orient='records')
    return []
appointments = load_appointments()

def save_appointments(appts):
    today_str = datetime.now().strftime('%d.%m.%Y')
    filtered = [a for a in appts if a['date'] != today_str]
    df = pd.DataFrame(filtered)
    df.to_excel(EXCEL_FILE, index=False)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    try:
        specialist_col = header.index("specialist") + 1
    except ValueError:
        specialist_col = None
    if specialist_col:
        for row in range(2, ws.max_row+1):
            cell = ws.cell(row=row, column=specialist_col)
            spec = cell.value
            color = specialist_colors.get(spec)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill
        wb.save(EXCEL_FILE)

# --- FSM на базе dict
user_states = {}
user_data = {}

def set_state(user_id, state):
    user_states[user_id] = state
def get_state(user_id):
    return user_states.get(user_id)
def update_data(user_id, **kwargs):
    user_data.setdefault(user_id, {}).update(kwargs)
def get_user_data(user_id):
    return user_data.get(user_id, {})

# --- Клавиатуры
from maxgram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton

def specialists_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=s)] for s in specialists],
        resize_keyboard=True, one_time_keyboard=True
    )
def massage_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=m)] for m in massage_types],
        resize_keyboard=True, one_time_keyboard=True
    )
def date_keyboard():
    today = datetime.now().date()
    keyboard = []
    for i in range(14):
        day = today + timedelta(days=i)
        keyboard.append([KeyboardButton(text=day.strftime("%d.%m.%Y"))])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, one_time_keyboard=True)

def time_keyboard_for_specialist(specialist, date_str):
    keyboard = []
    today = datetime.now().date()
    try:
        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
    except Exception:
        date_obj = today
    if specialist == "Соловьева":
        shift_start, shift_end = dt_time(8, 0), dt_time(13, 0)
    elif specialist == "Соловьев":
        shift_start, shift_end = dt_time(14, 0), dt_time(19, 0)
    else:
        shift_start, shift_end = work_start, work_end
    current = datetime.combine(datetime.today(), shift_start)
    end = datetime.combine(datetime.today(), shift_end)
    lunch_s = datetime.combine(datetime.today(), lunch_start)
    lunch_e = datetime.combine(datetime.today(), lunch_end)
    times = []
    while current < end:
        if not (lunch_s <= current < lunch_e):
            if date_obj == today and current.time() <= datetime.now().time():
                current += timedelta(minutes=30)
                continue
            times.append(current.strftime("%H:%M"))
        current += timedelta(minutes=30)
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t)] for t in times],
        resize_keyboard=True, one_time_keyboard=True
    )

def confirmation_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Подтвердить", callback_data="confirm"),
                InlineKeyboardButton(text="Отмена", callback_data="cancel"),
                InlineKeyboardButton(text="Меню", callback_data="menu"),
            ]
        ]
    )

def is_time_free(date_str, time_str):
    for appt in appointments:
        if appt["date"] == date_str and appt["time"] == time_str:
            return False
    return True

# --- Основная логика
@bot.message("start")
async def start_handler(msg):
    user_id = msg.from_user.id
    set_state(user_id, "waiting_for_specialist")
    await msg.answer("Выберите специалиста:", reply_markup=specialists_keyboard())

@bot.message()
async def handle(msg):
    user_id = msg.from_user.id
    state = get_state(user_id)
    text = msg.text.strip() if hasattr(msg, 'text') else ''
    if not state:
        return await msg.answer("Отправьте команду /start для начала.")
    if state == "waiting_for_specialist":
        if text in specialists:
            update_data(user_id, specialist=text)
            set_state(user_id, "waiting_for_massage")
            await msg.answer("Выберите вид массажа:", reply_markup=massage_keyboard())
        else:
            await msg.answer("Пожалуйста, выберите специалиста по кнопке.", reply_markup=specialists_keyboard())
    elif state == "waiting_for_massage":
        if text in massage_types:
            update_data(user_id, massage=text)
            set_state(user_id, "waiting_for_date")
            await msg.answer("Выберите дату:", reply_markup=date_keyboard())
        else:
            await msg.answer("Используйте кнопки для выбора вида массажа.", reply_markup=massage_keyboard())
    elif state == "waiting_for_date":
        try:
            date_obj = datetime.strptime(text, "%d.%m.%Y").date()
            if date_obj < datetime.now().date():
                raise ValueError
        except ValueError:
            return await msg.answer("Неверный формат даты!", reply_markup=date_keyboard())
        update_data(user_id, date=text)
        set_state(user_id, "waiting_for_time")
        specialist = get_user_data(user_id).get("specialist")
        await msg.answer("Выберите время:", reply_markup=time_keyboard_for_specialist(specialist, text))
    elif state == "waiting_for_time":
        specialist = get_user_data(user_id).get("specialist")
        try:
            datetime.strptime(text, "%H:%M")
        except ValueError:
            return await msg.answer("Неверный формат времени.", reply_markup=time_keyboard_for_specialist(specialist, get_user_data(user_id).get("date")))
        cur_date = get_user_data(user_id).get("date")
        if not is_time_free(cur_date, text):
            return await msg.answer("Это время занято.", reply_markup=time_keyboard_for_specialist(specialist, cur_date))
        update_data(user_id, time=text)
        set_state(user_id, "waiting_for_name")
        await msg.answer("Как к вам обращаться? Напишите имя или обращение:")
    elif state == "waiting_for_name":
        if len(text) < 2:
            return await msg.answer("Пожалуйста, введите корректное имя.")
        update_data(user_id, name=text)
        set_state(user_id, "waiting_for_phone")
        await msg.answer("Пожалуйста, введите ваш номер телефона (через +7...):")
    elif state == "waiting_for_phone":
        phone = text.replace(" ", "")
        if not (phone.startswith("+7") and len(phone) >= 11):
            return await msg.answer("Введите корректный номер телефона.")
        update_data(user_id, phone=phone)
        data = get_user_data(user_id)
        text_msg = (
            f"Проверьте данные записи:\n"
            f"Дата и время: {data['date']} {data['time']}\n"
            f"Специалист: {data['specialist']}\n"
            f"Вид массажа: {data['massage']}\n"
            f"Как обращаться: {data['name']}\n"
            f"Телефон: {data['phone']}\n\n"
            "Подтвердите запись."
        )
        set_state(user_id, "waiting_for_confirmation")
        await msg.answer(text_msg, reply_markup=confirmation_keyboard())

# --- Inline callback-обработка
@bot.callback("confirm")
async def confirm_callback(call):
    user_id = call.from_user.id
    data = get_user_data(user_id)
    new_appointment = {
        "user_id": user_id,
        "specialist": data["specialist"],
        "massage": data["massage"],
        "date": data["date"],
        "time": data["time"],
        "name": data["name"],
        "phone": data["phone"]
    }
    appointments.append(new_appointment)
    save_appointments(appointments)
    await call.message.edit_text("Запись подтверждена!")
    set_state(user_id, None)
    user_data.pop(user_id, None)

@bot.callback("cancel")
async def cancel_callback(call):
    user_id = call.from_user.id
    await call.message.edit_text("Запись отменена.")
    set_state(user_id, None)
    user_data.pop(user_id, None)

@bot.callback("menu")
async def menu_callback(call):
    user_id = call.from_user.id
    await call.message.edit_text("Возврат в главное меню.")
    set_state(user_id, None)
    user_data.pop(user_id, None)

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())


#pip install maxgram
#pip install pandas openpyxl apscheduler
