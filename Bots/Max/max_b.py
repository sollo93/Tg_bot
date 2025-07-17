import asyncio
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from os.path import exists
from datetime import datetime, timedelta, time as dt_time

from maxgram import MaxBot, MaxDispatcher, types
from maxgram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton

API_TOKEN = "ВАШ_MAX_ТОКЕН"
EXCEL_FILE = "appointments.xlsx"
OFFICIAL_SITE = "https://ваш-сайт.ру"
SOCIAL_LINKS = [
    ("Instagram", "https://instagram.com/your_profile"),
    ("VK", "https://vk.com/your_profile"),
]

bot = MaxBot(token=API_TOKEN)
dp = MaxDispatcher()

# --- Базовые данные ---

specialists = ["Соловьева", "Соловьев"]
massage_types = ["Классический", "Точечный", "Спортивный", "Расслабляющий", "Антицеллюлитный"]
work_start, work_end = dt_time(8, 0), dt_time(23, 0)
lunch_start, lunch_end = dt_time(13, 0), dt_time(14, 0)
specialist_colors = {"Соловьева": "6272A4", "Соловьев": "50FA7B"}
main_services = ["Массаж", "Парикмахер", "Ногти"]

def load_appointments():
    if exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
        return df.to_dict(orient='records')
    return []

appointments = load_appointments()

def save_appointments(appts):
    df = pd.DataFrame(appts)
    df.to_excel(EXCEL_FILE, index=False)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    try:
        specialist_col = header.index("specialist") + 1
    except ValueError:
        specialist_col = None
    if specialist_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=specialist_col)
            spec = cell.value
            color = specialist_colors.get(spec)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill
    wb.save(EXCEL_FILE)

user_states = {}
user_data = {}

def set_state(user_id, state):
    user_states[user_id] = state
def get_state(user_id):
    return user_states.get(user_id)
def update_data(user_id, **kwargs):
    user_data.setdefault(user_id, {}).update(kwargs)
def get_user_data(user_id):
    return user_data.get(user_id, {})

# --- Клавиатуры ---
def main_menu_keyboard():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📝 Запись на приём")],
        [KeyboardButton(text="🔖 Мои записи"), KeyboardButton(text="🌐 Оф. сайт")],
        [KeyboardButton(text="🌍 Соц.сети")]
    ], resize_keyboard=True)

def services_keyboard():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="💆 Массаж")],
        [KeyboardButton(text="💇 Парикмахер")],
        [KeyboardButton(text="💅 Ногти")],
        [KeyboardButton(text="◀️ Назад")]
    ], resize_keyboard=True)

def massage_specialists_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=s)] for s in specialists],
        resize_keyboard=True, one_time_keyboard=True)

def massage_types_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=m)] for m in massage_types],
        resize_keyboard=True, one_time_keyboard=True)

def date_keyboard():
    today = datetime.now().date()
    keyboard = [[KeyboardButton(text=(today + timedelta(days=i)).strftime("%d.%m.%Y"))] for i in range(14)]
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, one_time_keyboard=True)

def time_keyboard_for_specialist(specialist, date_str):
    today = datetime.now().date()
    try:
        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
    except Exception:
        date_obj = today
    if specialist == "Соловьева":
        shift_start, shift_end = dt_time(8, 0), dt_time(13, 0)
    elif specialist == "Соловьев":
        shift_start, shift_end = dt_time(14, 0), dt_time(19, 0)
    else:
        shift_start, shift_end = work_start, work_end
    current = datetime.combine(datetime.today(), shift_start)
    end = datetime.combine(datetime.today(), shift_end)
    lunch_s = datetime.combine(datetime.today(), lunch_start)
    lunch_e = datetime.combine(datetime.today(), lunch_end)
    times = []
    while current < end:
        if not (lunch_s <= current < lunch_e):
            if date_obj == today and current.time() <= datetime.now().time():
                current += timedelta(minutes=30)
                continue
            times.append(current.strftime("%H:%M"))
        current += timedelta(minutes=30)
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t)] for t in times],
        resize_keyboard=True, one_time_keyboard=True
    )

def confirmation_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Подтвердить", callback_data="confirm"),
                InlineKeyboardButton(text="Отмена", callback_data="cancel"),
                InlineKeyboardButton(text="Меню", callback_data="menu"),
            ]
        ]
    )

def records_keyboard(records):
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=f"🗑 Отменить: {record['date']} {record['service']} {record.get('specialist', '')}", callback_data=f"cancel_record:{record['id']}")]
            for record in records if not record.get("canceled")
        ]
    ) if records else None

def social_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=name, url=url)] for name, url in SOCIAL_LINKS]
    )

def is_time_free(date_str, time_str, specialist):
    for appt in appointments:
        if appt["date"] == date_str and appt["time"] == time_str and not appt.get("canceled") and appt["specialist"] == specialist:
            return False
    return True

# --- FSM и обработчики ---

@bot.message("start")
async def start_handler(msg):
    set_state(msg.from_user.id, "main_menu")
    await msg.answer("Добро пожаловать! Выберите действие:", reply_markup=main_menu_keyboard())

@bot.message()
async def handle_menu(msg):
    text, user_id = (msg.text or "").strip(), msg.from_user.id
    state = get_state(user_id) or "main_menu"

    # Главное меню
    if text in ("◀️ Назад", "/start"):
        set_state(user_id, "main_menu")
        await msg.answer("Главное меню:", reply_markup=main_menu_keyboard())

    elif state == "main_menu":
        if text == "📝 Запись на приём":
            set_state(user_id, "service_select")
            await msg.answer("Выберите направление:", reply_markup=services_keyboard())
        elif text == "🔖 Мои записи":
            user_records = [a for a in appointments if a.get("user_id") == user_id and not a.get("canceled")]
            if not user_records:
                await msg.answer("У вас нет активных записей.")
            else:
                await msg.answer("Ваши записи:", reply_markup=records_keyboard(user_records))
        elif text == "🌐 Оф. сайт":
            await msg.answer(f"Сайт: {OFFICIAL_SITE}")
        elif text == "🌍 Соц.сети":
            await msg.answer("Мы в соцсетях:", reply_markup=social_keyboard())
        else:
            await msg.answer("Пожалуйста, выберите действие через меню.")

    # Выбор услуги
    elif state == "service_select":
        if text == "💆 Массаж":
            set_state(user_id, "massage_specialist")
            await msg.answer("Выберите специалиста:", reply_markup=massage_specialists_keyboard())
        elif text == "💇 Парикмахер":
            await msg.answer("Онлайн-запись к парикмахеру появится в след.обновлении.")
        elif text == "💅 Ногти":
            await msg.answer("Онлайн-запись на ногти появится в след.обновлении.")
        else:
            await msg.answer("Пожалуйста, выберите доступный раздел.", reply_markup=services_keyboard())

    # ---- Массаж: ваш FSM-алгоритм ----
    elif state == "massage_specialist":
        if text in specialists:
            update_data(user_id, service="Массаж", specialist=text)
            set_state(user_id, "massage_type")
            await msg.answer("Выберите вид массажа:", reply_markup=massage_types_keyboard())
        else:
            await msg.answer("Пожалуйста, выберите специалиста кнопкой.", reply_markup=massage_specialists_keyboard())
    elif state == "massage_type":
        if text in massage_types:
            update_data(user_id, massage=text)
            set_state(user_id, "massage_date")
            await msg.answer("Выберите дату:", reply_markup=date_keyboard())
        else:
            await msg.answer("Пожалуйста, выберите вид массажа по кнопке.", reply_markup=massage_types_keyboard())
    elif state == "massage_date":
        try:
            datetime.strptime(text, "%d.%m.%Y")
            update_data(user_id, date=text)
            set_state(user_id, "massage_time")
            specialist = get_user_data(user_id).get("specialist")
            await msg.answer("Выберите время:", reply_markup=time_keyboard_for_specialist(specialist, text))
        except Exception:
            await msg.answer("Введите дату через кнопки.", reply_markup=date_keyboard())
    elif state == "massage_time":
        specialist = get_user_data(user_id).get("specialist")
        date = get_user_data(user_id).get("date")
        try:
            datetime.strptime(text, "%H:%M")
            if not is_time_free(date, text, specialist):
                await msg.answer("Время занято, выберите другое.", reply_markup=time_keyboard_for_specialist(specialist, date))
                return
            update_data(user_id, time=text)
            set_state(user_id, "massage_name")
            await msg.answer("Как к вам обращаться? Напишите имя или обращение:")
        except Exception:
            await msg.answer("Неверный формат времени.", reply_markup=time_keyboard_for_specialist(specialist, date))
    elif state == "massage_name":
        if len(text) >= 2:
            update_data(user_id, name=text)
            set_state(user_id, "massage_phone")
            await msg.answer("Введите номер телефона (через +7...):")
        else:
            await msg.answer("Пожалуйста, введите имя или обращение.")
    elif state == "massage_phone":
        phone = text.replace(" ", "")
        if not (phone.startswith("+7") and len(phone) >= 11):
            await msg.answer("Введите корректный номер телефона (пример: +79112223344).")
            return
        update_data(user_id, phone=phone)
        data = get_user_data(user_id)
        data["user_id"] = user_id
        data["service"] = "Массаж"
        data["id"] = f"{datetime.now().timestamp()}_{user_id}"
        set_state(user_id, "massage_confirm")
        appointment_text = (
            f"Детали записи:\n"
            f"Дата: {data['date']} {data['time']}\n"
            f"Специалист: {data['specialist']}\n"
            f"Массаж: {data['massage']}\n"
            f"Имя: {data['name']}\n"
            f"Телефон: {data['phone']}\n"
            "Подтвердите запись."
        )
        await msg.answer(appointment_text, reply_markup=confirmation_keyboard())
    elif state == "massage_confirm":
        await msg.answer("Подтвердите или отмените запись кнопками.", reply_markup=confirmation_keyboard())
    else:
        await msg.answer("Пожалуйста, используйте меню или кнопки.")

# --- Callback обработчики подтверждения и отмены ---
@bot.callback("confirm")
async def confirm_cb(call):
    user_id = call.from_user.id
    data = get_user_data(user_id)
    appointments.append({**data})
    save_appointments(appointments)
    await call.message.edit_text("Запись подтверждена!")
    set_state(user_id, "main_menu")
    user_data.pop(user_id, None)

@bot.callback("cancel")
async def cancel_cb(call):
    user_id = call.from_user.id
    await call.message.edit_text("Запись отменена.")
    set_state(user_id, "main_menu")
    user_data.pop(user_id, None)

@bot.callback("menu")
async def menu_cb(call):
    user_id = call.from_user.id
    await call.message.edit_text("Главное меню.", reply_markup=main_menu_keyboard())
    set_state(user_id, "main_menu")
    user_data.pop(user_id, None)

@bot.callback()
async def cancel_record_cb(call):
    if call.data.startswith("cancel_record:"):
        rec_id = call.data.split(":", 1)[1]
        found = False
        for appt in appointments:
            if appt["id"] == rec_id and appt.get("user_id") == call.from_user.id:
                appt["canceled"] = True
                found = True
                save_appointments(appointments)
                break
        if found:
            await call.message.edit_text("Запись отменена!")
        else:
            await call.message.edit_text("Запись не найдена.")

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
