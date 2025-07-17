import asyncio
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from os.path import exists
from datetime import datetime, timedelta, time as dt_time
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, ReplyKeyboardRemove
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler

API_TOKEN = '7537212499:AAEg5u2M21e6kQkQn-9D_ylQNtOml1EB3NE'
EXCEL_FILE = 'appointments.xlsx'


bot = Bot(token=API_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler()

async def main():
    scheduler.start()
    await dp.start_polling(bot)
    
# Данные
specialists = ["Соловьева", "Соловьев", "Сидоров"]
massage_types = ["Классический", "Точечный", "Спортивный", "Расслабляющий", "Антицеллюлитный"]
work_start = dt_time(8, 0)
work_end = dt_time(23, 0)
lunch_start = dt_time(13, 0)
lunch_end = dt_time(14, 0)

specialist_user_ids = {
    "Соловьева": 1152741301,
    "Соловьев": 812829316,
    # "Сидоров": 192837465
}
# Словарь цветов для специалистов (цвета в формате HEX без #)
specialist_colors = {
    #"Соловьева": "ADD8E6",  # светло-голубой
    "Соловьева": "6272A4",
    "Соловьев": "50FA7B",6
    #"Соловьев": "FFC7CE",   # светло-красный
    #"Сидоров": "C6EFCE"     # светло-зеленый
}
# Загрузка записей из Excel при старте
def load_appointments():6
    if exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
        return df.to_dict(orient='records')
    return []

#appointments = []
# Загрузка данных при старте
appointments = load_appointments()


# Добавляем новые состояния для обращения и телефона
class Booking(StatesGroup):
    waiting_for_specialist = State()
    waiting_for_massage = State()
    waiting_for_date = State()
    waiting_for_time = State()
    waiting_for_name = State()       # новое состояние для обращения
    waiting_for_phone = State()      # новое состояние для номера телефона
    waiting_for_confirmation = State()


# Сохранение записей в Excel (исключая записи на сегодняшний день)
def save_appointments(appts):
    today_str = datetime.now().strftime('%d.%m.%Y')
    filtered = [a for a in appts if a['date'] != today_str]
    df = pd.DataFrame(filtered)
    df.to_excel(EXCEL_FILE, index=False)

    # Открываем файл для форматирования
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Находим индекс столбца "Специалист"
    header = [cell.value for cell in ws[1]]
    try:
        specialist_col = header.index("specialist") + 1  # +1, т.к. openpyxl 1-индексация
    except ValueError:
        specialist_col = None

    if specialist_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=specialist_col)
            spec = cell.value
            color = specialist_colors.get(spec)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill

    wb.save(EXCEL_FILE)


def specialists_keyboard():
    builder = ReplyKeyboardBuilder()
    for s in specialists:
        builder.button(text=s)
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def massage_keyboard():
    builder = ReplyKeyboardBuilder()
    for m in massage_types:
        builder.button(text=m)
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def date_keyboard():
    builder = ReplyKeyboardBuilder()
    today = datetime.now().date()
    for i in range(14):
        day = today + timedelta(days=i)
        builder.button(text=day.strftime("%d.%m.%Y"))
    builder.adjust(3)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

# Изменяем функцию time_keyboard, чтобы учитывать смены специалистов
def time_keyboard_for_specialist(specialist, date_str):
    builder = ReplyKeyboardBuilder()
    times = []
    today = datetime.now().date()
    try:
        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
    except Exception:
        date_obj = today

    # Определяем смену по специалисту
    if specialist == "Соловьева":
        shift_start = dt_time(8, 0)
        shift_end = dt_time(13, 0)
    elif specialist == "Соловьев":
        shift_start = dt_time(14, 0)
        shift_end = dt_time(19, 0)
    else:
        shift_start = work_start
        shift_end = work_end

    current = datetime.combine(datetime.today(), shift_start)
    end = datetime.combine(datetime.today(), shift_end)
    lunch_s = datetime.combine(datetime.today(), lunch_start)
    lunch_e = datetime.combine(datetime.today(), lunch_end)

    while current < end:
        # Исключаем обеденный перерыв, если он внутри смены
        if not (lunch_s <= current < lunch_e):
            # Если дата — сегодня, исключаем время меньше текущего
            if date_obj == today and current.time() <= datetime.now().time():
                current += timedelta(minutes=30)
                continue
            times.append(current.strftime("%H:%M"))
        current += timedelta(minutes=30)

    for t in times:
        builder.button(text=t)
    builder.adjust(3)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def confirmation_keyboard():
    builder = InlineKeyboardBuilder()
    builder.button(text="Подтвердить", callback_data="confirm")
    builder.button(text="Отмена", callback_data="cancel")
    builder.button(text="Меню", callback_data="menu")
    builder.adjust(3)
    return builder.as_markup()

def is_time_free(date_str, time_str):
    for appt in appointments:
        if appt["date"] == date_str and appt["time"] == time_str:
            return False
    return True

async def send_reminder(user_id: int, text: str):
    try:
        await bot.send_message(user_id, text)
    except Exception as e:
        print(f"Ошибка при отправке уведомления {user_id}: {e}")

@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    await message.answer("Выберите специалиста:", reply_markup=specialists_keyboard())
    await state.set_state(Booking.waiting_for_specialist)

@dp.message(Booking.waiting_for_specialist, F.text.in_(specialists))
async def specialist_chosen(message: Message, state: FSMContext):
    await state.update_data(specialist=message.text)
    await message.answer(f"Специалист {message.text} выбран.\nВыберите вид массажа:", reply_markup=massage_keyboard())
    await state.set_state(Booking.waiting_for_massage)

@dp.message(Booking.waiting_for_massage, F.text.in_(massage_types))
async def massage_chosen(message: Message, state: FSMContext):
    await state.update_data(massage=message.text)
    await message.answer("Выберите дату для записи:", reply_markup=date_keyboard())
    await state.set_state(Booking.waiting_for_date)

@dp.message(Booking.waiting_for_date)
async def date_chosen(message: Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text, "%d.%m.%Y").date()
        if date_obj < datetime.now().date():
            await message.answer("Дата не может быть в прошлом. Пожалуйста, выберите корректную дату.", reply_markup=date_keyboard())
            return
    except ValueError:
        await message.answer("Неверный формат даты. Используйте ДД.MM.ГГГГ.", reply_markup=date_keyboard())
        return
    await state.update_data(date=message.text)
    data = await state.get_data()
    specialist = data.get("specialist")
    await message.answer("Выберите время:", reply_markup=time_keyboard_for_specialist(specialist, message.text))
    await state.set_state(Booking.waiting_for_time)


@dp.message(Booking.waiting_for_time)
async def time_chosen(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text, "%H:%M")
    except ValueError:
        await message.answer("Неверный формат времени. Используйте ЧЧ:ММ.", reply_markup=time_keyboard())
        return
    data = await state.get_data()
    date = data.get("date")
    specialist = data.get("specialist")
    time_ = message.text

    # Проверяем занятость времени
    if not is_time_free(date, time_):
        await message.answer("Это время уже занято, выберите другое.", reply_markup=time_keyboard_for_specialist(specialist, date))
        return

    await state.update_data(time=time_)
    await message.answer("Пожалуйста, укажите, как к вам обращаться (имя или обращение):", reply_markup=ReplyKeyboardRemove())
    await state.set_state(Booking.waiting_for_name)

@dp.callback_query(F.data == "confirm", Booking.waiting_for_confirmation)
async def confirm_booking(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = callback.from_user.id
    new_appointment = {
        "user_id": user_id,
        "specialist": data["specialist"],
        "massage": data["massage"],
        "date": data["date"],
        "time": data["time"],
        "name": data["name"],
        "phone": data["phone"]
    }
    appointments.append(new_appointment)
    save_appointments(appointments)  # Сохраняем в Excel

    dt_str = f"{data['date']} {data['time']}"
    dt_obj = datetime.strptime(dt_str, "%d.%m.%Y %H:%M")
    reminder_time = dt_obj - timedelta(hours=1)
    if reminder_time > datetime.now():
        scheduler.add_job(send_reminder, 'date', run_date=reminder_time,
                          args=[user_id, f"Напоминание: запись к {data['specialist']} ({data['name']}) на массаж '{data['massage']}' через 1 час ({dt_str})."])
    await callback.message.edit_text(
        "Запись подтверждена!\n\n"
        f"Дата и время: {data['date']} {data['time']}\n"
        f"Специалист: {data['specialist']}\n"
        f"Вид массажа: {data['massage']}\n"
        f"Как обращаться: {data['name']}\n"
        f"Телефон: {data['phone']}"
    )
    await state.clear()
    await cmd_start(callback.message, state)


@dp.message(Booking.waiting_for_name)
async def name_chosen(message: Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer("Пожалуйста, введите корректное имя или обращение.")
        return
    await state.update_data(name=name)
    # Запрос номера телефона с клавиатурой для отправки контакта
    keyboard = ReplyKeyboardBuilder()
    keyboard.button(text="Отправить номер телефона", request_contact=True)
    keyboard.adjust(1)
    await message.answer("Пожалуйста, отправьте ваш номер телефона:", reply_markup=keyboard.as_markup(resize_keyboard=True, one_time_keyboard=True))
    await state.set_state(Booking.waiting_for_phone)

@dp.message(Booking.waiting_for_phone, F.contact)
async def phone_chosen_contact(message: Message, state: FSMContext):
    phone = message.contact.phone_number
    await state.update_data(phone=phone)
    await confirm_data(message, state)

@dp.message(Booking.waiting_for_phone)
async def phone_chosen_text(message: Message, state: FSMContext):
    phone = message.text.strip()
    if not phone or len(phone) < 5:
        await message.answer("Пожалуйста, введите корректный номер телефона или отправьте контакт.")
        return
    await state.update_data(phone=phone)
    await confirm_data(message, state)

async def confirm_data(message: Message, state: FSMContext):
    data = await state.get_data()
    text = (
        f"Проверьте данные записи:\n"
        f"Дата и время: {data['date']} {data['time']}\n"
        f"Специалист: {data['specialist']}\n"
        f"Вид массажа: {data['massage']}\n"
        f"Как обращаться: {data['name']}\n"
        f"Телефон: {data['phone']}\n\n"
        f"Подтвердите запись."
    )
    await message.answer(text, reply_markup=confirmation_keyboard())
    await state.set_state(Booking.waiting_for_confirmation)

@dp.callback_query(F.data == "confirm", Booking.waiting_for_confirmation)
async def confirm_booking(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = callback.from_user.id
    appointments.append({
        #"user_id": user_id,
        "Специалист": data["Специалист"],
        "Вид массажа": data["Вид массажа"],
        "Дата и время": data["Дата и время"],
        "Имя": data["Имя"],
        "Телефон": data["Телефон"]
    })
    dt_str = f"{data['date']} {data['time']}"
    dt_obj = datetime.strptime(dt_str, "%d.%m.%Y %H:%M")
    reminder_time = dt_obj - timedelta(hours=1)
    if reminder_time > datetime.now():
        scheduler.add_job(send_reminder, 'date', run_date=reminder_time,
                          args=[user_id, f"Напоминание: запись к {data['specialist']} ({data['name']}) на массаж '{data['massage']}' через 1 час ({dt_str})."])
    await callback.message.edit_text(
        "Запись подтверждена!\n\n"
        f"Дата и время: {data['date']} {data['time']}\n"
        f"Специалист: {data['specialist']}\n"
        f"Вид массажа: {data['massage']}\n"
        f"Как обращаться: {data['name']}\n"
        f"Телефон: {data['phone']}"
    )
    await state.clear()
    await cmd_start(callback.message, state)

@dp.callback_query(F.data == "cancel", Booking.waiting_for_confirmation)
async def cancel_booking(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Запись отменена.")
    await state.clear()
    await cmd_start(callback.message, state)

@dp.callback_query(F.data == "menu")
async def menu_callback(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Возврат в главное меню.")
    await state.clear()
    await cmd_start(callback.message, state)

@dp.message(Command("list"))
async def my_appointments(message: Message):
    user_id = message.from_user.id
    filtered = [a for a in appointments if specialist_user_ids.get(a["specialist"]) == user_id]
    if not filtered:
        await message.answer("У вас пока нет записей.")
        return
    text = "Ваши записи:\n\n"
    for appt in filtered:
        text += (
            f"{appt['date']} {appt['time']}\n"
            f"Специалист: {appt['specialist']}\n"
            f"Вид массажа: {appt['massage']}\n"
            f"Как обращаться: {appt.get('name', 'Не указано')}\n"
            f"Телефон: {appt.get('phone', 'Не указан')}\n\n"
        )
    await message.answer(text)

@dp.message(Command("myid"))
async def send_user_id(message: Message):
    user_id = message.from_user.id
    await message.answer(f"Ваш user ID: {user_id}")

@dp.message()
async def fallback(message: Message):
    await message.answer("Пожалуйста, используйте кнопки для выбора.")

if __name__ == "__main__":
    asyncio.run(main())
