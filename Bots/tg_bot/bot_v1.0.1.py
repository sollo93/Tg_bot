import asyncio
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from os.path import exists
from datetime import datetime, timedelta, time as dt_time
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import (
    Message, CallbackQuery, ReplyKeyboardRemove,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler

main_menu_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="оф.сайт")],
        [KeyboardButton(text="запись на прием")],
        [KeyboardButton(text="Мои записи")],
        [KeyboardButton(text="наши соц.сети")]
    ],
    resize_keyboard=True
)

API_TOKEN = '7537212499:AAEg5u2M21e6kQkQn-9D_ylQNtOml1EB3NE'  # Безопаснее убрать токен из открытого кода.
EXCEL_FILE = 'appointments.xlsx'

bot = Bot(token=API_TOKEN)
dp = Dispatcher()
scheduler = AsyncIOScheduler()

async def main():
    scheduler.start()
    await dp.start_polling(bot)

specialists = ["Соловьева", "Соловьев"]
massage_types = ["Классический", "Точечный", "Спортивный", "Расслабляющий", "Антицеллюлитный"]
work_start = dt_time(8, 0)
work_end = dt_time(23, 0)
lunch_start = dt_time(13, 0)
lunch_end = dt_time(14, 0)

specialist_user_ids = {
    "Соловьева": 1152741301,
    "Соловьев": 812829316,
}
specialist_colors = {
    "Соловьева": "6272A4",
    "Соловьев": "50FA7B",
}
def load_appointments():
    if exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
        return df.to_dict(orient='records')
    return []

appointments = load_appointments()

class Booking(StatesGroup):
    waiting_for_specialist = State()
    waiting_for_massage = State()
    waiting_for_date = State()
    waiting_for_time = State()
    waiting_for_name = State()
    waiting_for_phone = State()
    waiting_for_confirmation = State()

def save_appointments(appts):
    today_str = datetime.now().strftime('%d.%m.%Y')
    filtered = [a for a in appts if a['date'] != today_str]
    df = pd.DataFrame(filtered)
    df.to_excel(EXCEL_FILE, index=False)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    try:
        specialist_col = header.index("specialist") + 1
    except ValueError:
        specialist_col = None
    if specialist_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=specialist_col)
            spec = cell.value
            color = specialist_colors.get(spec)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill
    wb.save(EXCEL_FILE)


def my_appointments_inline(user_id):
    kb = InlineKeyboardBuilder()
    user_appts = [a for a in appointments if a['user_id'] == user_id]
    for idx, appt in enumerate(user_appts):
        text = f"{appt['date']} {appt['time']} к {appt['specialist']}"
        kb.button(
            text=f"Отменить {text}",
            callback_data=f"cancel_{idx}"
        )
    kb.adjust(1)
    return kb.as_markup() if user_appts else None

@dp.message(F.text == "Мои записи")
async def show_user_appts(message: Message):
    user_id = message.from_user.id
    kb = my_appointments_inline(user_id)
    if not kb:
        await message.answer("У вас пока нет записей.")
        return
    await message.answer("Ваши записи:\n\n(нажмите 'Отменить' для удаления записи)", reply_markup=kb)

@dp.callback_query(F.data.startswith("cancel_"))
async def cancel_user_appt(callback: CallbackQuery):
    user_id = callback.from_user.id
    idx = int(callback.data.split("_")[1])
    user_appts = [a for a in appointments if a['user_id'] == user_id]
    if idx < 0 or idx >= len(user_appts):
        await callback.answer("Запись не найдена.")
        return
    to_remove = user_appts[idx]
    appointments.remove(to_remove)
    save_appointments(appointments)
    await callback.answer("Запись отменена.", show_alert=True)
    kb = my_appointments_inline(user_id)
    if kb:
        await callback.message.edit_text("Обновленный список записей:", reply_markup=kb)
    else:
        await callback.message.edit_text("Все ваши записи отменены.")

def date_keyboard(specialist=None):
    builder = ReplyKeyboardBuilder()
    today = datetime.now().date()
    for i in range(14):
        day = today + timedelta(days=i)
        date_str = day.strftime("%d.%m.%Y")
        free_slot_exists = False
        if specialist == "Соловьева":
            shift_start, shift_end = dt_time(8, 0), dt_time(13, 0)
        elif specialist == "Соловьев":
            shift_start, shift_end = dt_time(14, 0), dt_time(19, 0)
        else:
            shift_start, shift_end = work_start, work_end
        current = datetime.combine(day, shift_start)
        end = datetime.combine(day, shift_end)
        while current < end:
            t = current.strftime("%H:%M")
            if is_time_free(date_str, t, specialist):
                free_slot_exists = True
                break
            current += timedelta(minutes=30)
        if free_slot_exists:
            builder.button(text=date_str)
    builder.adjust(3)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def is_time_free(date_str, time_str, specialist=None):
    return not any(
        a["date"] == date_str and a["time"] == time_str and (specialist is None or a["specialist"] == specialist)
        for a in appointments
    )

def time_keyboard_for_specialist(specialist, date_str):
    builder = ReplyKeyboardBuilder()
    today = datetime.now().date()
    try:
        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
    except Exception:
        date_obj = today
    if specialist == "Соловьева":
        shift_start, shift_end = dt_time(8, 0), dt_time(13, 0)
    elif specialist == "Соловьев":
        shift_start, shift_end = dt_time(14, 0), dt_time(19, 0)
    else:
        shift_start, shift_end = work_start, work_end
    current = datetime.combine(date_obj, shift_start)
    end = datetime.combine(date_obj, shift_end)
    lunch_s = datetime.combine(date_obj, lunch_start)
    lunch_e = datetime.combine(date_obj, lunch_end)
    while current < end:
        t_str = current.strftime("%H:%M")
        if (not (lunch_s <= current < lunch_e)) and is_time_free(date_str, t_str, specialist):
            if not (date_obj == today and current.time() <= datetime.now().time()):
                builder.button(text=t_str)
        current += timedelta(minutes=30)
    builder.adjust(3)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def specialists_keyboard():
    builder = ReplyKeyboardBuilder()
    for s in specialists:
        builder.button(text=s)
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def massage_keyboard():
    builder = ReplyKeyboardBuilder()
    for m in massage_types:
        builder.button(text=m)
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

def confirmation_keyboard():
    builder = InlineKeyboardBuilder()
    builder.button(text="Подтвердить", callback_data="confirm")
    builder.button(text="Отмена", callback_data="cancel")
    builder.button(text="Меню", callback_data="menu")
    builder.adjust(3)
    return builder.as_markup()

async def send_reminder(user_id: int, text: str):
    try:
        await bot.send_message(user_id, text)
    except Exception as e:
        print(f"Ошибка при отправке уведомления {user_id}: {e}")

@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Главное меню:", reply_markup=main_menu_keyboard)

@dp.message(F.text=="запись на прием")
async def booking_menu(message: Message, state: FSMContext):
    await message.answer("Выберите специалиста:", reply_markup=specialists_keyboard())
    await state.set_state(Booking.waiting_for_specialist)

site_button = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="Перейти на сайт", url="https://example.com")]
    ]
)

@dp.message(F.text=="оф.сайт")
async def official_site(message: Message):
    await message.answer("Официальный сайт: [укажите ссылку здесь]", reply_markup=site_button)

@dp.message(F.text=="наши соц.сети")
async def social_networks(message: Message):
    await message.answer("Наши соц.сети: [укажите ссылки здесь]")

@dp.message(Booking.waiting_for_specialist, F.text.in_(specialists))
async def specialist_chosen(message: Message, state: FSMContext):
    await state.update_data(specialist=message.text)
    await message.answer(f"Специалист {message.text} выбран.\nВыберите вид массажа:", reply_markup=massage_keyboard())
    await state.set_state(Booking.waiting_for_massage)

@dp.message(Booking.waiting_for_massage, F.text.in_(massage_types))
async def massage_chosen(message: Message, state: FSMContext):
    data = await state.get_data()
    specialist = data.get("specialist")
    await state.update_data(massage=message.text)
    # Ошибка 3: Было date_keyboard() без параметра, теперь правильно:
    await message.answer("Выберите дату для записи:", reply_markup=date_keyboard(specialist))
    await state.set_state(Booking.waiting_for_date)

@dp.message(Booking.waiting_for_date)
async def date_chosen(message: Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text, "%d.%m.%Y").date()
        if date_obj < datetime.now().date():
            data = await state.get_data()
            specialist = data.get("specialist")
            await message.answer(
                "Дата не может быть в прошлом. Пожалуйста, выберите корректную дату.",
                reply_markup=date_keyboard(specialist)
            )
            return
    except ValueError:
        data = await state.get_data()
        specialist = data.get("specialist")
        await message.answer(
            "Неверный формат даты. Используйте ДД.MM.ГГГГ.",
            reply_markup=date_keyboard(specialist)
        )
        return
    await state.update_data(date=message.text)
    data = await state.get_data()
    specialist = data.get("specialist")
    await message.answer(
        "Выберите время:",
        reply_markup=time_keyboard_for_specialist(specialist, message.text)
    )
    await state.set_state(Booking.waiting_for_time)

@dp.message(Booking.waiting_for_time)
async def time_chosen(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text, "%H:%M")
    except ValueError:
        data = await state.get_data()
        specialist = data.get("specialist")
        date = data.get("date")
        await message.answer(
            "Неверный формат времени. Используйте ЧЧ:ММ.",
            reply_markup=time_keyboard_for_specialist(specialist, date)
        )
        return
    data = await state.get_data()
    date = data.get("date")
    specialist = data.get("specialist")
    time_ = message.text
    if not is_time_free(date, time_, specialist):
        await message.answer(
            "Это время уже занято, выберите другое.",
            reply_markup=time_keyboard_for_specialist(specialist, date)
        )
        return
    await state.update_data(time=time_)
    await message.answer(
        "Пожалуйста, укажите, как к вам обращаться (имя или обращение):",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(Booking.waiting_for_name)

@dp.callback_query(F.data == "confirm", Booking.waiting_for_confirmation)
async def confirm_booking(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user_id = callback.from_user.id
    new_appointment = {
        "user_id": user_id,
        "specialist": data["specialist"],
        "massage": data["massage"],
        "date": data["date"],
        "time": data["time"],
        "name": data["name"],
        "phone": data["phone"]
    }
    appointments.append(new_appointment)
    save_appointments(appointments)
    dt_str = f"{data['date']} {data['time']}"
    dt_obj = datetime.strptime(dt_str, "%d.%m.%Y %H:%M")
    reminder_time = dt_obj - timedelta(hours=1)
    if reminder_time > datetime.now():
        scheduler.add_job(
            send_reminder, 'date', run_date=reminder_time,
            args=[user_id,
                  f"Напоминание: запись к {data['specialist']} ({data['name']}) на массаж '{data['massage']}' через 1 час ({dt_str})."]
        )
    await callback.message.edit_text(
        "Запись подтверждена!\n\n"
        f"Дата и время: {data['date']} {data['time']}\n"
        f"Специалист: {data['specialist']}\n"
        f"Вид массажа: {data['massage']}\n"
        f"Как обращаться: {data['name']}\n"
        f"Телефон: {data['phone']}"
    )
    await state.clear()
    await cmd_start(callback.message, state)

@dp.message(Booking.waiting_for_name)
async def name_chosen(message: Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer("Пожалуйста, введите корректное имя или обращение.")
        return
    await state.update_data(name=name)
    keyboard = ReplyKeyboardBuilder()
    keyboard.button(text="Отправить номер телефона", request_contact=True)
    keyboard.adjust(1)
    await message.answer("Пожалуйста, отправьте ваш номер телефона:", reply_markup=keyboard.as_markup(resize_keyboard=True, one_time_keyboard=True))
    await state.set_state(Booking.waiting_for_phone)

@dp.message(Booking.waiting_for_phone, F.contact)
async def phone_chosen_contact(message: Message, state: FSMContext):
    phone = message.contact.phone_number
    await state.update_data(phone=phone)
    await confirm_data(message, state)

@dp.message(Booking.waiting_for_phone)
async def phone_chosen_text(message: Message, state: FSMContext):
    phone = message.text.strip()
    if not phone or len(phone) < 5:
        await message.answer("Пожалуйста, введите корректный номер телефона или отправьте контакт.")
        return
    await state.update_data(phone=phone)
    await confirm_data(message, state)

async def confirm_data(message: Message, state: FSMContext):
    data = await state.get_data()
    text = (
        f"Проверьте данные записи:\n"
        f"Дата и время: {data['date']} {data['time']}\n"
        f"Специалист: {data['specialist']}\n"
        f"Вид массажа: {data['massage']}\n"
        f"Как обращаться: {data['name']}\n"
        f"Телефон: {data['phone']}\n\n"
        f"Подтвердите запись."
    )
    await message.answer(text, reply_markup=confirmation_keyboard())
    await state.set_state(Booking.waiting_for_confirmation)

@dp.callback_query(F.data == "cancel", Booking.waiting_for_confirmation)
async def cancel_booking(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Запись отменена.")
    await state.clear()
    await cmd_start(callback.message, state)

@dp.callback_query(F.data == "menu")
async def menu_callback(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Возврат в главное меню.")
    await state.clear()
    await cmd_start(callback.message, state)

@dp.message(Command("list"))
async def my_appointments(message: Message):
    user_id = message.from_user.id
    filtered = [a for a in appointments if specialist_user_ids.get(a["specialist"]) == user_id]
    if not filtered:
        await message.answer("У вас пока нет записей.")
        return
    text = "Ваши записи:\n\n"
    for appt in filtered:
        text += (
            f"{appt['date']} {appt['time']}\n"
            f"Специалист: {appt['specialist']}\n"
            f"Вид массажа: {appt['massage']}\n"
            f"Как обращаться: {appt.get('name', 'Не указано')}\n"
            f"Телефон: {appt.get('phone', 'Не указан')}\n\n"
        )
    await message.answer(text)

@dp.message(Command("myid"))
async def send_user_id(message: Message):
    user_id = message.from_user.id
    await message.answer(f"Ваш user ID: {user_id}")

@dp.message()
async def fallback(message: Message):
    await message.answer("Пожалуйста, используйте кнопки для выбора.")

if __name__ == "__main__":
    asyncio.run(main())
